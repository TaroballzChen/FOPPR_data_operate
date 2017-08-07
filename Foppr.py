import os
import glob
import xlsxwriter
import openpyxl
import xlrd
from win32com.client import Dispatch

def data_transfer_to_list(origin_file_name):                         #將原始lvm data 記錄到列表中
    global transfer_data
    transfer_data = []
    with open(str(origin_file_name), 'r')as origin_data:
        data = origin_data.readlines()
        for each_data in data:
            data_per_sec = each_data.rstrip('\n').split('\t')
            transfer_data.append(data_per_sec)


def calc_formula(filename="get_point&average.xlsx"):                           #開啟excel檔案並存檔後關閉excel
    xls = Dispatch("Excel.Application")
    return xls.Workbooks.Open(os.path.abspath(os.getcwd()+"\\"+filename)).Close(True)

def create_xlsx(filename="originData.xlsx"):                                             #創建excel檔紀錄原始數據
    os.chdir(os.getcwd())                                                           #移動到當前資料夾路徑
    All_of_file = (glob.glob('*.lvm'))                                              #顯示當前資料夾.lvm檔案
    data_workbook  = xlsxwriter.Workbook(filename)                                  #創建xlsx
    for origin_filename in All_of_file:
        data_worksheet = data_workbook.add_worksheet('%s'%origin_filename)          #將檔名依序寫入sheet中
        data_transfer_to_list(origin_filename)

        def data_write():                                                           #將原始數據寫入每個sheet中 A列為實驗秒數 B列為signal
            sec_and_signal = 1
            for each_data in transfer_data:
                data_worksheet.write('A%d'%sec_and_signal,float(each_data[0]))
                data_worksheet.write('B%d'%sec_and_signal,float(each_data[1]))
                sec_and_signal += 1

        data_write()

    data_workbook.close() #存檔


def data_operate(filename="originData.xlsx",total_select_input=int(input('請輸入每一濃度的取點數:')),get_point_input=int(input("請輸入共幾個濃度(包含blank):"))):
    global select_list_All
    select_list_All=[]                                        #創建一個表蒐集各個sheet的實驗
    all_data = openpyxl.load_workbook(filename)               #開啟xlsx檔案
    exp_sheet_names = all_data.get_sheet_names()              #列出所有sheet(每次實驗)名   //列表
    global get_point
    global total_select
    total_select = total_select_input                         # 一個濃度的訊號取幾點平均     //使用者輸入
    total_select_record = total_select
    get_point = get_point_input                               # 共取幾個濃度作圖(包含blank)  //使用者輸入

    for each_sheetname in exp_sheet_names:                    #開始對sheet操作
        ws = all_data.get_sheet_by_name(each_sheetname)

        point_of_x = []
        for _input in range(get_point):
            if _input == 0:
                _input="blank"
            user_input = int(input("%s_"%each_sheetname+"Point_of_x_%s:"%str(_input)))     #取點的x座標(會以此點往前推算使用者輸入的取點數)   //使用者輸入
            point_of_x.append(user_input)
        else:
            print("此次實驗取點分別為",point_of_x)
        i = 0                                             #point_of_x函數為第幾項
        exp_list = []                                     #創建一個列表蒐集所有sheet的實驗
        for each_select_point in point_of_x:
            select_list = []                              #創建一個表蒐集但一濃度不同點(初始化)
            row = point_of_x[i]

            while total_select != 0:
                select_point = ws.cell(row=row+1,column=2)  #從選擇的點開始往前取點
                select_list.append(select_point.value)      #將所選的值紀錄
                total_select -= 1
                row -=1
            exp_list.append(select_list)                    #將單一濃度的所有取點加到列表中
            total_select = total_select_record              #一個濃度取幾點平均 //恢復使用者設定
            i+=1                                            #進行同次實驗下一個濃度取點
        select_list_All.append(exp_list)                    #將各個實驗各個濃度的取點記錄到列表中
    all_data.close()

def operated_data_write(filename="get_point&average.xlsx"):
    os.chdir(os.getcwd())                                   # 移動到當前資料夾路徑
    All_of_file = (glob.glob('*.lvm'))                      # 讀取.lvm檔案
    data_workbook = xlsxwriter.Workbook(filename)           # 創建xlsx
    bold = data_workbook.add_format({'bold': True})         # 設置粗體格式

    global concentration
    blank = 'blank'
    concentration = []  # 濃度  //使用者輸入(低濃度高濃度)
    for num in range(get_point-1):
        concentration_input=input(str("請由低到高依序輸入濃度(EXCEL格式):"))
        concentration.append(concentration_input)
    concentration.insert(0,blank)
    print ('各個濃度為:',concentration)

    num_exp = 0
    for origin_filename in All_of_file:
        data_worksheet = data_workbook.add_worksheet('%s' % origin_filename)      # 將檔名依序寫入sheet中
        for_conc = 0

        for conc in concentration:                                                #寫各個濃度的標題
            data_worksheet.write(chr(65 + for_conc) + '1', conc, bold)
            data_worksheet.write(chr(65 + for_conc)+ '2', "=ABS(LOG10(" +conc+ "))")
            write_start = 3

            for each_point_per_con in select_list_All[num_exp][for_conc]:
                data_worksheet.write_number(chr(65 + for_conc) + str(write_start), each_point_per_con)
                write_start += 1
            else:                                                              #迴圈完成則後計算平均
                def avgsignal_per_con():
                    global cal_avg
                    cal_avg=write_start+2
                    data_worksheet.write(chr(65+for_conc)+str(cal_avg),"=ROUND(AVERAGE("+chr(65+for_conc)+str(write_start-total_select)+":"+chr(65+for_conc)+str(write_start)+"),7)")
                avgsignal_per_con()
            for_conc+=1
        num_exp+=1
    data_workbook.close()


def read_average_and_logconc(filename="get_point&average.xlsx"):
    global total_avg_value
    global total_logconc_value
    total_avg_value=[]
    total_logconc_value=[]
    operated_data=openpyxl.load_workbook(filename,data_only=True)
    all_sheet_name=operated_data.get_sheet_names()

    for sheet_name in all_sheet_name:
        each_sheet = operated_data.get_sheet_by_name(sheet_name)
        for_conc = 1
        sheet_all_avg_value = []
        sheet_all_logconc = []
        for i in range(get_point):
            avg_value = each_sheet.cell(row=cal_avg,column=for_conc).value
            log_conc  = each_sheet.cell(row=2,column=for_conc).value
            sheet_all_avg_value.append(avg_value)
            sheet_all_logconc.append(log_conc)
            for_conc+=1
        total_avg_value.append(sheet_all_avg_value)
        del sheet_all_logconc[0]
        total_logconc_value.append(sheet_all_logconc)
    print("各濃度由低到高 取點後訊號平均值各為：",total_avg_value)                     #印出所有濃度signal平均值
    print("各濃度由低到高 取log值後分別為：",total_logconc_value)                     #印出所有取完log的濃度值

def read_avg(filename="get_point&average.xlsx"):
    global total_avg_value2
    total_avg_value2=[]
    operated_data=xlrd.open_workbook(filename)
    all_sheet_name=operated_data.sheet_names()
    for sheet_name in all_sheet_name:
        each_sheet = operated_data.sheet_by_name(sheet_name)
        for_conc=0
        sheet_all_avg_value=[]
        for i in range(get_point):
            avg_value = each_sheet.cell(cal_avg,for_conc)
            sheet_all_avg_value.append(avg_value)
            for_conc+=1
        total_avg_value2.append(sheet_all_avg_value)
    print(total_avg_value2)

def caculate_normalize_conc():
    global nomalize_signal_result
    nomalize_signal_result=[]
    number_exp=0
    for num_exp in total_avg_value:
        num_conc=len(total_avg_value[total_avg_value.index(num_exp)])-1
        each_exp_normalize_data=[]
        for avg_data in  total_avg_value[number_exp]:
            nomalize_avg_data = float(total_avg_value[number_exp][num_conc]-total_avg_value[number_exp][0])/total_avg_value[number_exp][num_conc]   #I-I0/I
            each_exp_normalize_data.insert(0,nomalize_avg_data)
            num_conc-=1
            if num_conc == 0:
                nomalize_signal_result.append(each_exp_normalize_data)
                break
            else:
                pass
        number_exp += 1
    print("訊號由低到高訊號normalize的結果分別為：",nomalize_signal_result)

def create_result_data():
    All_of_file = glob.glob("*.lvm")
    j = 0
    for each_file in All_of_file:
        i = 0
        with open(each_file+'_.txt','w') as result:
            if len(nomalize_signal_result[j]) == len(total_logconc_value[j]):
                for i in range(len(nomalize_signal_result[j])):
                    result.write(str(total_logconc_value[j][i])+'\t'+str(nomalize_signal_result[j][i])+'\n')

                    i+=1
            else:
                print("normalize sigal & log cocetartion doesn't match with the list is out of range")
        j+=1

create_xlsx()
data_operate()
operated_data_write()
calc_formula()
read_average_and_logconc()
caculate_normalize_conc()
create_result_data()
print ('-'*20,'Designed by Yuan-Yu Chen','-'*20)
print('-'*20,'有問題請提出討論','-'*20)